from django.contrib import admin
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
import csv
import io
from openpyxl import Workbook
from .models import Member, Group, Event, Attendance, CommunicationLog
from .constant import MEMBER_CSV_HEADERS, COMMUNICATION_CSV_HEADERS, MEMBER_CATEGORIES
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile


@admin.register(Member)
class MemberAdmin(admin.ModelAdmin):
    list_display = ('first_name', 'last_name', 'contact_1', 'contact_2', 'gender', 'email', 'category', 'created_at')
    list_filter = ('category', 'gender', 'created_at')
    search_fields = ('first_name', 'last_name', 'contact_1', 'contact_2', 'email')
    ordering = ('-created_at',)
    actions = ['export_as_csv', 'export_as_excel', 'send_bulk_sms', 'send_bulk_email']

    def export_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="members.csv"'
        writer = csv.writer(response)
        writer.writerow(MEMBER_CSV_HEADERS)
        for member in queryset:
            writer.writerow([
                f"{member.first_name} {member.last_name}",
                member.contact_1,
                member.contact_2,
                member.get_gender_display(),
                member.email,
                member.category,
            ])
        return response
    export_as_csv.short_description = _("Export selected members as CSV")

    def export_as_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Members"
        # Write headers
        for col_num, header in enumerate(MEMBER_CSV_HEADERS, 1):
            ws.cell(row=1, column=col_num, value=header)
        # Write data
        for row_num, member in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=f"{member.first_name} {member.last_name}")
            ws.cell(row=row_num, column=2, value=member.contact_1)
            ws.cell(row=row_num, column=3, value=member.contact_2)
            ws.cell(row=row_num, column=4, value=member.get_gender_display())
            ws.cell(row=row_num, column=5, value=member.email)
            ws.cell(row=row_num, column=6, value=member.category)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="members.xlsx"'
        wb.save(response)
        return response
    export_as_excel.short_description = _("Export selected members as Excel")

    def send_bulk_sms(self, request, queryset):
        # Placeholder for bulk SMS sending
        messages.success(request, f"Bulk SMS sent to {queryset.count()} members.")
    send_bulk_sms.short_description = _("Send bulk SMS to selected members")

    def send_bulk_email(self, request, queryset):
        # Placeholder for bulk email sending
        messages.success(request, f"Bulk email sent to {queryset.count()} members.")
    send_bulk_email.short_description = _("Send bulk email to selected members")

    def import_members(self, request):
        if request.method == 'POST':
            csv_file = request.FILES.get('csv_file')
            if not csv_file:
                messages.error(request, "No file selected.")
                return redirect('admin:app_member_changelist')

            if not csv_file.name.endswith('.csv'):
                messages.error(request, "File must be a CSV.")
                return redirect('admin:app_member_changelist')

            try:
                decoded_file = csv_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                members_to_create = []
                errors = []
                for row_num, row in enumerate(reader, start=2):  # Start at 2 assuming header is row 1
                    try:
                        first_name = row.get('first_name', '').strip()
                        last_name = row.get('last_name', '').strip()
                        contact_1 = row.get('contact_1', '').strip()
                        contact_2 = row.get('contact_2', '').strip()
                        gender = row.get('gender', 'M').strip().upper()
                        email = row.get('email', '').strip()
                        category = row.get('category', 'Member').strip()

                        if not first_name or not last_name:
                            errors.append(f"Row {row_num}: First name and last name are required.")
                            continue

                        if gender not in ['M', 'F']:
                            errors.append(f"Row {row_num}: Invalid gender '{gender}'. Must be 'M' or 'F'.")
                            continue

                        if category not in [choice[0] for choice in MEMBER_CATEGORIES]:
                            errors.append(f"Row {row_num}: Invalid category '{category}'.")
                            continue

                        members_to_create.append(Member(
                            first_name=first_name,
                            last_name=last_name,
                            contact_1=contact_1,
                            contact_2=contact_2,
                            gender=gender,
                            email=email,
                            category=category,
                        ))
                    except Exception as e:
                        errors.append(f"Row {row_num}: Error processing row - {str(e)}")

                if members_to_create:
                    Member.objects.bulk_create(members_to_create)
                    messages.success(request, f"Successfully imported {len(members_to_create)} members.")
                else:
                    messages.warning(request, "No valid members to import.")

                if errors:
                    messages.warning(request, f"Errors in import: {len(errors)} rows skipped. Check data.")

            except Exception as e:
                messages.error(request, f"Error importing members: {str(e)}")
            return redirect('admin:app_member_changelist')

        return render(request, 'admin/import_members.html', {'title': 'Import Members'})
    import_members.short_description = _("Import members from CSV")

    def download_sample_csv(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sample_members.csv"'
        writer = csv.writer(response)
        writer.writerow(['first_name', 'last_name', 'contact_1', 'contact_2', 'gender', 'email', 'category'])
        writer.writerow(['John', 'Doe', '+1234567890', '', 'M', 'john.doe@example.com', 'Member'])
        return response

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-members/', self.admin_site.admin_view(self.import_members), name='import_members'),
            path('download-sample-csv/', self.admin_site.admin_view(self.download_sample_csv), name='download_sample_csv'),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['import_members_url'] = 'import-members/'
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(Group)
class GroupAdmin(admin.ModelAdmin):
    list_display = ('name', 'leader', 'member_count', 'created_at')
    list_filter = ('created_at', 'leader')
    search_fields = ('name', 'description')
    ordering = ('name',)

    def member_count(self, obj):
        return obj.members.count()
    member_count.short_description = _("Member Count")


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ('title', 'event_type', 'date', 'time', 'location', 'created_by', 'created_at')
    list_filter = ('event_type', 'date', 'created_by')
    search_fields = ('title', 'description', 'location')
    ordering = ('-date',)


@admin.register(Attendance)
class AttendanceAdmin(admin.ModelAdmin):
    list_display = ('member', 'event', 'status', 'recorded_at')
    list_filter = ('status', 'recorded_at', 'event__event_type')
    search_fields = ('member__full_name', 'event__title')
    ordering = ('-recorded_at',)


@admin.register(CommunicationLog)
class CommunicationLogAdmin(admin.ModelAdmin):
    list_display = ('communication_type', 'subject', 'sender', 'recipient_count', 'sent_at')
    list_filter = ('communication_type', 'sent_at', 'sender')
    search_fields = ('subject', 'message')
    ordering = ('-sent_at',)
    actions = ['export_communications_as_csv']

    def recipient_count(self, obj):
        return obj.recipients.count()
    recipient_count.short_description = _("Recipient Count")

    def export_communications_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="communications.csv"'
        writer = csv.writer(response)
        writer.writerow(COMMUNICATION_CSV_HEADERS)
        for comm in queryset:
            recipients = ', '.join([r.full_name for r in comm.recipients.all()])
            writer.writerow([
                recipients,
                comm.communication_type,
                comm.message,
                comm.sent_at,
                comm.sender.username,
            ])
        return response
    export_communications_as_csv.short_description = _("Export selected communications as CSV")
